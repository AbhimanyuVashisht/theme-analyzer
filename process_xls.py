from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Theming_Raw_Data.xlsx', read_only=True)
ws = wb.active
f = open('theme.txt', 'w+')
for k, row in enumerate(ws.rows):
    if k != 0:
        for i, cell in enumerate(row):
            if i % 2 == 0:
                review = ws[get_column_letter(i + 1) + str(k + 1)].value
                theme = ws[get_column_letter(i + 2) + str(k + 2)].value
                if review is not None and theme is not None:
                    line = str(review) + '\t' + str(theme) + '\n'
                    f.writelines(line)

f.close()

